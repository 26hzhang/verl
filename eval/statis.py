import os
import re
import pandas as pd
from collections import defaultdict

# ANSI颜色代码
class Colors:
    GREEN = '\033[92m'
    BOLD = '\033[1m'
    END = '\033[0m'
    HIGHLIGHT = '\033[93m'  # 黄色

def parse_model_path(path):
    """从model路径中提取xxx和yyy"""
    # 尝试匹配 verl_checkpoints 路径
    pattern1 = r'verl_checkpoints/(?:[^/]+/)?([^/]+)/global_step_(\d+)/'
    match = re.search(pattern1, path)
    if match:
        return match.group(1), int(match.group(2))
    # import pdb; pdb.set_trace()
    # 尝试匹配 models/xxx 路径
    pattern2 = r'models/([^/]+)'
    match = re.search(pattern2, path)
    if match:
        return match.group(1), 0
    
    pattern3 = r'Models/([^/]+)'
    match = re.search(pattern3, path)
    if match:
        return match.group(1), 0
    
    
    return None, None

def check_threshold(task_name, metric_name, value):
    """检查是否满足阈值条件"""
    if pd.isna(value):
        return False
    
    try:
        value = float(value)
    except (ValueError, TypeError):
        return False
    # print(task_name)
    # 定义阈值规则
    thresholds = {
        'aime2024': 0.6,
        'aime2025': 0.366,
        'amc23': 0.9,
        'gpqa': 0.7,
        'math500': 0.888,
        'math': 0.925,
        
    }
    
    # 检查任务名是否包含关键字（不区分大小写）
    task_lower = task_name.lower()
    for key, threshold in thresholds.items():
        if key.lower() in task_lower:
            return value >= threshold
    
    return False

def highlight_value(value, is_highlighted):
    """为值添加高亮标记"""
    if is_highlighted:
        return f"{Colors.GREEN}{Colors.BOLD}{value}{Colors.END}"
    return str(value)

def process_csv_files(directory):
    """处理目录中的所有CSV文件"""
    # 存储数据: {xxx: {yyy: {任务名: {指标: 值}}}}
    data_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    
    csv_files = [f for f in os.listdir(directory) if f.endswith('pass_k.csv')]
    
    for csv_file in csv_files:
        file_path = os.path.join(directory, csv_file)
        task_name = csv_file.replace('.csv', '')
        
        try:
            df = pd.read_csv(file_path)
            
            for idx, row in df.iterrows():
                model_path = row['model']
                xxx, yyy = parse_model_path(model_path)
                
                if xxx is not None and yyy is not None:
                    for col in df.columns:
                        if col != 'model':
                            data_dict[xxx][yyy][task_name][col] = row[col]
        
        except Exception as e:
            print(f"处理文件 {csv_file} 时出错: {e}")
    # import pdb; pdb.set_trace()
    return data_dict

def create_task_grouped_report(xxx, step_data):
    """为一个模型创建按任务分组的报告（带高亮）"""
    all_tasks = set()
    for yyy, tasks in step_data.items():
        all_tasks.update(tasks.keys())
    
    all_tasks = sorted(all_tasks)
    
    report_lines = []
    report_lines.append(f"\n{'='*100}")
    report_lines.append(f"模型: {xxx}")
    report_lines.append(f"{'='*100}\n")
    
    # 为每个任务创建一个表格
    for task_name in all_tasks:
        report_lines.append(f"\n任务: {task_name}")
        report_lines.append("-" * 100)
        
        # 收集该任务的数据
        rows = []
        highlighted_info = []  # 记录高亮信息
        
        for yyy in sorted(step_data.keys()):
            if task_name in step_data[yyy]:
                row = {'step': yyy}
                row.update(step_data[yyy][task_name])
                rows.append(row)
        
        if rows:
            df = pd.DataFrame(rows)
            df = df.sort_values('step').reset_index(drop=True)
            
            # 创建高亮版本的DataFrame用于显示
            df_display = df.copy()
            for col in df.columns:
                if col != 'step':
                    df_display[col] = df[col].apply(
                        lambda x: highlight_value(x, check_threshold(task_name, col, x))
                    )
            
            # 格式化输出
            report_lines.append(df_display.to_string(index=False))
            
            # 添加注释说明
            has_highlight = any(
                check_threshold(task_name, col, val)
                for col in df.columns if col != 'step'
                for val in df[col]
            )
            # if has_highlight:
            #     report_lines.append(f"\n{Colors.GREEN}★ = 超过阈值的优秀结果{Colors.END}")
            
            report_lines.append("")
    
    return "\n".join(report_lines)

def create_compact_csv_report(xxx, step_data):
    """创建紧凑的CSV格式报告（所有任务在一个表格中，但列名简短）"""
    all_rows = []
    
    for yyy in sorted(step_data.keys()):
        row = {'step': yyy}
        
        for task_name, metrics in step_data[yyy].items():
            for metric, value in metrics.items():
                col_name = f"{task_name[:15]}_{metric}"
                row[col_name] = value
        
        all_rows.append(row)
    
    df = pd.DataFrame(all_rows)
    df = df.sort_values('step').reset_index(drop=True)
    
    return df

def create_excel_report(xxx, step_data, output_file):
    """创建Excel报告，每个任务一个sheet（带条件格式）"""
    try:
        from openpyxl.styles import PatternFill, Font
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            all_tasks = set()
            for yyy, tasks in step_data.items():
                all_tasks.update(tasks.keys())
            
            for task_name in sorted(all_tasks):
                rows = []
                for yyy in sorted(step_data.keys()):
                    if task_name in step_data[yyy]:
                        row = {'step': yyy}
                        row.update(step_data[yyy][task_name])
                        rows.append(row)
                
                if rows:
                    df = pd.DataFrame(rows)
                    df = df.sort_values('step').reset_index(drop=True)
                    
                    sheet_name = task_name[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 添加条件格式（高亮满足条件的单元格）
                    worksheet = writer.sheets[sheet_name]
                    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    highlight_font = Font(bold=True, color='006100')
                    
                    for row_idx, row_data in enumerate(rows, start=2):
                        for col_idx, col_name in enumerate(df.columns, start=1):
                            if col_name != 'step':
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                if check_threshold(task_name, col_name, cell.value):
                                    cell.fill = highlight_fill
                                    cell.font = highlight_font
        
        return True
    except Exception as e:
        print(f"创建Excel文件失败: {e}")
        return False

# 使用示例
if __name__ == "__main__":
    directory = "./"
    
    print("正在处理CSV文件...")
    data_dict = process_csv_files(directory)
    
    print(f"\n找到 {len(data_dict)} 个模型\n")
    print(f"阈值设置:")
    print(f"  - aime2024 > 0.5667")
    print(f"  - aime2025 > 0.3")
    print(f"  - AMC2023 > 0.85")
    print(f"  - GPQA > 0.6")
    
    # 为每个模型生成报告
    for xxx, step_data in data_dict.items():
        # 1. 打印按任务分组的美化报告（带高亮）
        report = create_task_grouped_report(xxx, step_data)
        print(report)
        

        
        print()